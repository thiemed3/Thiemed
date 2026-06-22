import base64
import io

from odoo.tests.common import TransactionCase
from odoo.exceptions import UserError
from odoo.tests import Form


class TestHomologationImport(TransactionCase):
    """Test Excel import wizard."""

    def setUp(self):
        super().setUp()
        # Ensure base_unit_count column has a default (website_sale adds it as required)
        self.env.cr.execute(
            "ALTER TABLE product_template ALTER COLUMN base_unit_count SET DEFAULT 1"
        )
        self.env.cr.execute(
            "ALTER TABLE product_product ALTER COLUMN base_unit_count SET DEFAULT 1"
        )
        self.partner = self.env["res.partner"].create({
            "name": "Competitor Test",
            "competitor": True,
        })
        self.product_a = self.env["product.product"].create({
            "name": "Pinza Disección 12cm",
            "default_code": "PIN-001",
        })
        self.product_b = self.env["product.product"].create({
            "name": "Portaagujas Mayo 14cm",
            "default_code": "POR-001",
        })

    def _make_excel(self, rows):
        """Create an Excel file in memory with header + data rows."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Código", "Descripción", "Código interno"])
        for row in rows:
            ws.append(row)
        out = io.BytesIO()
        wb.save(out)
        return base64.b64encode(out.getvalue())

    def test_01_import_matches_by_internal_code(self):
        """Rows with matching internal_code get validated."""
        excel = self._make_excel([
            ("CLI-001", "Pinza diseccion", "PIN-001"),
        ])
        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        wiz.action_import()

        h = self.env["product.homologation"].search([
            ("customer_code", "=", "CLI-001"),
        ])
        self.assertTrue(h)
        self.assertEqual(h.product_id, self.product_a)
        self.assertEqual(h.state, "validated")

    def test_02_import_matches_by_customer_code(self):
        """When customer_code == default_code, it should match."""
        excel = self._make_excel([
            ("PIN-001", "Pinza diseccion", ""),
        ])
        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        wiz.action_import()

        h = self.env["product.homologation"].search([
            ("customer_code", "=", "PIN-001"),
        ])
        self.assertTrue(h)
        self.assertEqual(h.product_id, self.product_a)
        self.assertEqual(h.precision_pct, 100.0)

    def test_03_import_no_match_creates_draft(self):
        """Rows with no match create draft with empty product."""
        excel = self._make_excel([
            ("UNKN-001", "Unknown product", ""),
        ])
        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        wiz.action_import()

        h = self.env["product.homologation"].search([
            ("customer_code", "=", "UNKN-001"),
        ])
        self.assertTrue(h)
        self.assertFalse(h.product_id)
        self.assertEqual(h.state, "draft")

    def test_04_import_empty_file_raises(self):
        """Empty file should raise UserError."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        wb = openpyxl.Workbook()
        out = io.BytesIO()
        wb.save(out)
        excel = base64.b64encode(out.getvalue())

        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        with self.assertRaises(UserError):
            wiz.action_import()

    def test_05_import_missing_code_column(self):
        """Excel without a code column raises UserError."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Precio"])
        ws.append(["Test", "100"])
        out = io.BytesIO()
        wb.save(out)
        excel = base64.b64encode(out.getvalue())

        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        with self.assertRaises(UserError):
            wiz.action_import()

    def test_06_import_multiple_rows(self):
        """Multiple rows with mixed match status are handled."""
        excel = self._make_excel([
            ("PIN-001", "Pinza exacta", ""),
            ("UNKN-002", "Sin match", ""),
            ("POR-001", "Portaagujas", ""),
        ])
        wiz = Form(self.env["product.homologation.import.wizard"])
        wiz.competitor_id = self.partner
        wiz.file = excel
        wiz = wiz.save()
        wiz.action_import()

        self.assertEqual(wiz.imported_count, 3)
        self.assertEqual(wiz.matched_count, 2)
        self.assertEqual(wiz.unmatched_count, 1)

    def test_07_form_ui_import_wizard(self):
        """Uses Form to simulate full import UI flow."""
        excel = self._make_excel([
            ("POR-001", "Portaagujas", ""),
        ])
        form = Form(self.env["product.homologation.import.wizard"])
        form.competitor_id = self.partner
        form.file = excel
        wiz = form.save()
        wiz.action_import()
        self.assertEqual(wiz.state, "result")
        self.assertIn("match exacto", wiz.result_summary.lower())
